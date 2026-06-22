from __future__ import annotations

from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
import csv
import io
import json
import re
from typing import Any

from openpyxl import load_workbook

from services.buying_intent_agent import analyze_buying_intent
from services.universal_engine import build_discovery_plan, build_universal_profile

try:  # pragma: no cover - optional dependency guard
    from services.ai_service import AIService
except Exception:  # noqa: BLE001
    AIService = None


SERVICE_ALIAS_BANK: list[tuple[str, list[str], list[str]]] = [
    ("digital marketing", ["digital marketing", "marketing agency", "growth agency", "marketing partner"], ["marketing agency", "marketing partner", "external marketing team"]),
    ("seo", ["seo", "search engine optimization", "organic growth", "local seo", "technical seo"], ["seo agency", "seo help", "seo consultant", "search visibility help"]),
    ("paid growth", ["paid ads", "google ads", "meta ads", "ppc", "performance marketing", "roas"], ["paid ads agency", "google ads agency", "ppc agency", "performance marketing partner", "better roas"]),
    ("lead gen", ["lead generation", "demand gen", "pipeline", "prospecting", "appointment setting"], ["lead generation agency", "more qualified leads", "pipeline acceleration", "demand generation agency"]),
    ("content", ["content marketing", "content strategy", "ghostwriter", "thought leadership", "linkedin content"], ["content marketing agency", "content strategy help", "linkedin ghostwriter", "personal brand support"]),
    ("social", ["social media", "linkedin marketing", "community", "creator"], ["social media agency", "social media help", "linkedin marketing help"]),
    ("email", ["email marketing", "lifecycle", "automation sequence", "newsletter"], ["email marketing agency", "email automation help", "lifecycle email support"]),
    ("analytics", ["analytics", "reporting", "tracking", "attribution", "dashboard"], ["analytics consultant", "tracking setup help", "attribution help", "dashboard setup"]),
    ("revops", ["revops", "hubspot", "crm", "sales ops", "marketing ops", "lifecycle setup"], ["revops consultant", "hubspot cleanup", "crm setup help", "marketing ops support"]),
    ("web", ["website", "web design", "landing page", "wordpress", "cms", "site redesign"], ["web design agency", "website redesign help", "landing page agency", "wordpress development partner"]),
    ("shopify", ["shopify", "ecommerce", "e-commerce", "d2c", "store", "cart"], ["shopify agency", "shopify development help", "ecommerce growth agency", "d2c marketing agency"]),
    ("software", ["software development", "custom software", "saas", "platform", "app build", "mobile app"], ["software development partner", "custom software agency", "saas development support", "app development agency"]),
    ("ai automation", ["ai automation", "automation", "workflow", "operations", "implementation"], ["ai automation partner", "workflow automation help", "automation support", "ai implementation partner"]),
    ("ai agents", ["ai agents", "copilot", "internal ai tools", "agentic"], ["ai agent developer", "copilot builder", "internal ai tools", "ai agent implementation"]),
    ("support automation", ["customer support", "cx", "chatbot", "helpdesk", "support automation"], ["customer support automation", "chatbot developer", "support automation partner", "ai customer support help"]),
    ("recruitment automation", ["recruitment automation", "hiring workflow", "ai recruiting", "rpo"], ["recruitment automation", "hiring workflow automation", "ai recruiting help"]),
    ("youtube", ["youtube", "video marketing", "video growth"], ["youtube marketing agency", "youtube growth help", "video strategy support"]),
    ("brand", ["brand", "branding", "rebrand", "positioning", "executive branding"], ["brand strategy agency", "branding agency", "rebrand support", "thought leadership help"]),
    ("crisis", ["crisis", "reputation", "brand recovery", "comms"], ["crisis communication help", "reputation management agency", "brand recovery support"]),
    ("real estate", ["real estate", "property", "brokerage", "listing"], ["real estate marketing agency", "property marketing partner", "real estate lead generation"]),
    ("b2b", ["b2b", "enterprise", "abm", "account based", "procurement"], ["b2b marketing agency", "account based marketing support", "enterprise demand gen"]),
    ("gtm", ["gtm", "go to market", "market expansion", "launch"], ["gtm consultant", "go to market help", "market expansion support"]),
    ("strategy", ["strategy", "roadmap", "planning", "positioning"], ["digital strategy consultant", "growth roadmap help", "marketing roadmap"]),
]

UNIVERSAL_BUYER_SEEDS = [
    "agency recommendations",
    "marketing partner",
    "digital marketing partner",
    "external support",
    "trusted agency",
    "consultant to help us",
    "partner who can execute",
    "team to handle this",
    "help getting more customers",
    "outsourced support",
]

BUYER_PATTERNS = [
    "looking for {term}",
    "need {term}",
    "need help with {term}",
    "looking for {term} partner",
    "need {term} partner",
    "anyone recommend {term}",
    "anyone know a good {term}",
    "looking to outsource {term}",
    "need external {term}",
]

NEGATIVE_FILTERS = [
    "hiring",
    "we're hiring",
    "we are hiring",
    "job opening",
    "open role",
    "send resume",
    "apply now",
    "years of experience",
    "engineer",
    "developer",
    "analyst",
    "we help",
    "i help",
    "book a call",
    "dm me",
    "available for work",
    "open for projects",
    "free consultation",
    "how to",
    "tips for",
    "lessons learned",
    "best practices",
    "case study",
    "podcast",
    "webinar",
    "newsletter",
    "thought leadership",
    "the future of",
    "top 10",
    "guide",
    "tutorial",
    "trend",
]


@dataclass
class LeadVaultSpec:
    tenant_id: str
    company_name: str
    website: str
    services: list[str]
    geography: str
    icp: str
    positioning: str
    customer_examples: list[str]
    founder_profile: str
    target_audience: str
    notes: str
    profile: dict[str, Any]
    discovery_plan: dict[str, Any]
    buyer_clusters: list[str]
    linkedin_queries: list[str]
    google_queries: list[str]
    buyer_phrases: list[str]
    negative_filters: list[str]
    query_scores: list[dict[str, Any]] = field(default_factory=list)
    buying_intent_preview: dict[str, Any] | None = None
    python_snippet: str = ""
    generated_at_utc: str = ""
    generation_mode: str = "deterministic"


def _split_values(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        items = value
    else:
        items = re.split(r"[\n;,|]+", str(value))
    cleaned: list[str] = []
    seen: set[str] = set()
    for item in items:
        text = str(item).strip()
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(text)
    return cleaned


def _normalize_text(*values: Any) -> str:
    return " ".join(str(v) for v in values if v not in (None, "")).lower()


def _service_seeds(services: list[str], icp: str, positioning: str, customer_examples: list[str], founder_profile: str, notes: str) -> list[str]:
    corpus = _normalize_text(services, icp, positioning, customer_examples, founder_profile, notes)
    seeds: list[str] = []
    for _, aliases, family_seeds in SERVICE_ALIAS_BANK:
        if any(alias in corpus for alias in aliases):
            seeds.extend(family_seeds)
    return list(dict.fromkeys(seeds or UNIVERSAL_BUYER_SEEDS))


def _build_clusters(services: list[str], icp: str, positioning: str, customer_examples: list[str], founder_profile: str, notes: str) -> list[str]:
    corpus = _normalize_text(services, icp, positioning, customer_examples, founder_profile, notes)
    clusters: list[str] = []
    for family, aliases, _ in SERVICE_ALIAS_BANK:
        if any(alias in corpus for alias in aliases):
            clusters.extend([
                f"{family} buyer intent",
                f"{family} agency search",
                f"{family} partner search",
            ])
    clusters.extend([
        "agency recommendation",
        "consultant search",
        "vendor evaluation",
        "implementation partner",
        "outsourcing signal",
    ])
    return sorted(dict.fromkeys(clusters or ["general agency search", "buyer procurement intent"]))


def _score_phrase(phrase: str, company_name: str, website: str, services: list[str]) -> dict[str, Any]:
    result = analyze_buying_intent(
        phrase,
        account={
            "company": company_name,
            "website": website,
            "industry": " ".join(services),
            "source": "leadvault",
        },
    )
    return {
        "phrase": phrase,
        "score": result.intent_score,
        "intent": result.buying_stage,
        "signal_type": result.signal_type,
    }


def _clean_query_phrase(phrase: str) -> str:
    cleaned = re.sub(r"\b(partner|agency|consultant|support|help|team)\s+\1\b", r"\1", phrase, flags=re.I)
    cleaned = re.sub(r"\bexternal\s+external\b", "external", cleaned, flags=re.I)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def _candidate_phrases(seed: str) -> list[str]:
    phrases: list[str] = []
    lower = seed.lower()
    for pattern in BUYER_PATTERNS:
        if pattern.endswith(" partner") and lower.endswith(("partner", "agency", "consultant", "team")):
            continue
        if pattern.startswith("need external") and lower.startswith("external "):
            continue
        phrases.append(_clean_query_phrase(pattern.format(term=seed)))
    return list(dict.fromkeys(phrases))


def _build_phrase_bank(service_seeds: list[str], company_name: str, website: str, services: list[str]) -> tuple[list[str], list[str], list[str], list[dict[str, Any]]]:
    scores: list[dict[str, Any]] = []
    phrases: list[str] = []
    for seed in list(dict.fromkeys([*service_seeds, *UNIVERSAL_BUYER_SEEDS])):
        for phrase in _candidate_phrases(seed):
            phrases.append(phrase)
            scores.append(_score_phrase(phrase, company_name, website, services))

    ranked = sorted(
        {item["phrase"]: item for item in scores}.values(),
        key=lambda item: (-int(item["score"]), item["phrase"]),
    )
    buyer_phrases = [item["phrase"] for item in ranked[:200]]
    linkedin_queries = buyer_phrases[:80]
    google_queries = [f'site:linkedin.com/posts "{phrase}"' for phrase in buyer_phrases[:40]]
    return buyer_phrases, linkedin_queries, google_queries, ranked


def _render_python_snippet(spec: LeadVaultSpec) -> str:
    def list_block(values: list[str]) -> str:
        return "[\n" + ",\n".join(f"    {json.dumps(value)}" for value in values) + "\n]"

    return "\n".join([
        "# LeadVault generated query bank",
        f"# Company: {spec.company_name or 'N/A'}",
        f"# Generated: {spec.generated_at_utc}",
        "",
        f"BUYER_INTENT_CLUSTERS = {list_block(spec.buyer_clusters)}",
        "",
        f"LINKEDIN_QUERIES = {list_block(spec.linkedin_queries)}",
        "",
        f"SERP_QUERIES = {list_block(spec.google_queries)}",
        "",
        f"_HIRE_PHRASES = {list_block(spec.buyer_phrases)}",
        "",
        f"HARD_NEGATIVE_FILTERS = {list_block(spec.negative_filters)}",
    ])


def build_leadvault_spec(
    *,
    company_name: str = "",
    website: str = "",
    services: list[str] | None = None,
    geography: str = "",
    icp: str = "",
    positioning: str = "",
    customer_examples: list[str] | None = None,
    founder_profile: str = "",
    target_audience: str = "",
    notes: str = "",
    tenant_id: str = "default",
) -> LeadVaultSpec:
    services_list = _split_values(services or [])
    examples_list = _split_values(customer_examples or [])
    profile = build_universal_profile(
        company_name=company_name,
        website=website,
        services=services_list,
        geography=geography,
        icp=icp,
        positioning=positioning,
        customer_examples=examples_list,
        founder_profile=founder_profile,
    )
    discovery_plan = build_discovery_plan(
        company_name=company_name,
        website=website,
        services=services_list,
        geography=geography,
        icp=icp,
        positioning=positioning,
        customer_examples=examples_list,
        founder_profile=founder_profile,
    )
    seeds = _service_seeds(services_list, icp, positioning, examples_list, founder_profile, notes)
    buyer_phrases, linkedin_queries, google_queries, query_scores = _build_phrase_bank(seeds, company_name, website, services_list)
    buying_intent_preview = analyze_buying_intent(
        " ".join([company_name, website, icp, positioning, target_audience, notes, " ".join(services_list)]),
        account={"company": company_name, "website": website, "industry": " ".join(services_list), "source": "leadvault"},
    )
    spec = LeadVaultSpec(
        tenant_id=tenant_id,
        company_name=company_name,
        website=website,
        services=services_list,
        geography=geography,
        icp=icp,
        positioning=positioning,
        customer_examples=examples_list,
        founder_profile=founder_profile,
        target_audience=target_audience,
        notes=notes,
        profile=profile.__dict__,
        discovery_plan=discovery_plan.__dict__,
        buyer_clusters=_build_clusters(services_list, icp, positioning, examples_list, founder_profile, notes),
        linkedin_queries=linkedin_queries,
        google_queries=google_queries,
        buyer_phrases=buyer_phrases,
        negative_filters=list(dict.fromkeys(NEGATIVE_FILTERS)),
        query_scores=query_scores,
        buying_intent_preview=buying_intent_preview.__dict__,
        generated_at_utc=datetime.now(timezone.utc).isoformat(),
    )
    spec.python_snippet = _render_python_snippet(spec)
    return spec


async def build_leadvault_spec_with_llm(spec: LeadVaultSpec) -> LeadVaultSpec:
    if AIService is None:
        return spec
    try:
        ai = AIService()
        context = {
            "company_name": spec.company_name,
            "website": spec.website,
            "services": spec.services,
            "icp": spec.icp,
            "positioning": spec.positioning,
            "target_audience": spec.target_audience,
            "buyer_clusters": spec.buyer_clusters,
            "linkedin_queries": spec.linkedin_queries[:40],
            "google_queries": spec.google_queries[:20],
        }
        response = await ai.complete(
            messages=[{
                "role": "user",
                "content": (
                    "Refine this LeadVault buyer-intent query bank. Keep phrases short, customer-written, "
                    "and focused on external agency/partner/vendor demand. Return strict JSON with keys "
                    "buyer_clusters, linkedin_queries, google_queries, buyer_phrases, negative_filters. "
                    f"Context: {json.dumps(context, ensure_ascii=False)}"
                ),
            }],
            model="claude-3-5-haiku-20241022",
            max_tokens=1400,
            temperature=0.2,
        )
        data = json.loads(response)
        if isinstance(data, dict):
            spec.buyer_clusters = _split_values(data.get("buyer_clusters")) or spec.buyer_clusters
            spec.linkedin_queries = (_split_values(data.get("linkedin_queries")) or spec.linkedin_queries)[:80]
            spec.google_queries = (_split_values(data.get("google_queries")) or spec.google_queries)[:40]
            spec.buyer_phrases = (_split_values(data.get("buyer_phrases")) or spec.buyer_phrases)[:200]
            spec.negative_filters = _split_values(data.get("negative_filters")) or spec.negative_filters
            spec.generation_mode = "llm_enriched"
            spec.python_snippet = _render_python_snippet(spec)
    except Exception:
        return spec
    return spec


def parse_icp_rows(data: bytes, filename: str) -> list[dict[str, Any]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        text = data.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        return [{str(k).strip().lower(): v for k, v in row.items()} for row in reader]
    if lower.endswith((".xlsx", ".xlsm", ".xls")):
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        return [{header: value for header, value in zip(headers, values)} for values in rows[1:]]
    raise ValueError("Only .csv, .xlsx, or .xlsm files are supported")


def _row_value(row: dict[str, Any], aliases: list[str]) -> str:
    normalized = {str(key).strip().lower(): value for key, value in row.items()}
    for alias in aliases:
        value = normalized.get(alias)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def aggregate_icp_profile(rows: list[dict[str, Any]]) -> dict[str, Any]:
    aliases = {
        "company_name": ["company_name", "company", "account", "organization", "organisation", "client", "brand"],
        "website": ["website", "company website", "url", "site", "web"],
        "services": ["services", "service", "offerings", "solutions", "what we do", "core services"],
        "geography": ["geography", "geo", "region", "location", "market", "country"],
        "icp": ["icp", "ideal customer profile", "target customer", "target audience", "audience"],
        "positioning": ["positioning", "value proposition", "position", "about", "summary"],
        "customer_examples": ["customer examples", "examples", "clients", "case studies", "references"],
        "founder_profile": ["founder profile", "founder", "bio", "about founder"],
        "notes": ["notes", "details", "additional info", "observations", "requirements"],
    }
    collected: dict[str, list[str]] = {key: [] for key in aliases}
    for row in rows:
        for key, names in aliases.items():
            value = _row_value(row, names)
            if not value:
                continue
            if key in {"services", "customer_examples", "notes"}:
                collected[key].extend(_split_values(value))
            else:
                collected[key].append(value)
    return {
        "company_name": next((item for item in collected["company_name"] if item), ""),
        "website": next((item for item in collected["website"] if item), ""),
        "services": _split_values(collected["services"]),
        "geography": next((item for item in collected["geography"] if item), ""),
        "icp": " | ".join(dict.fromkeys(collected["icp"])),
        "positioning": " | ".join(dict.fromkeys(collected["positioning"])),
        "customer_examples": _split_values(collected["customer_examples"]),
        "founder_profile": " | ".join(dict.fromkeys(collected["founder_profile"])),
        "notes": " | ".join(dict.fromkeys(collected["notes"])),
        "row_count": len(rows),
    }


def leadvault_spec_as_dict(spec: LeadVaultSpec) -> dict[str, Any]:
    return asdict(spec)
