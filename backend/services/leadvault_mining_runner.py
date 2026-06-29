from __future__ import annotations

import csv
import html
import io
import re
import urllib.parse
import urllib.request
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from typing import Any

from openpyxl import load_workbook

from config import settings
from services.accepted_export_store import save_accepted_export
from services.buying_intent_agent import analyze_buying_intent
from services.final_verifier import verify_post
from services.leadvault_agent import build_leadvault_spec, leadvault_spec_as_dict
from services.rejected_audit_store import save_rejected_audit


BUYER_SENTENCE = re.compile(
    r"[^.!?\n]*(looking for|seeking|need|recommend|rfp|rfq|vendor evaluation|shortlist|outsourc|implementation partner)[^.!?\n]*[.!?]?",
    re.IGNORECASE,
)
TAG_RE = re.compile(r"<[^>]+>")
LINK_RE = re.compile(r'<a[^>]+class="result__a"[^>]+href="([^"]+)"[^>]*>(.*?)</a>', re.IGNORECASE | re.DOTALL)
LINKEDIN_TEXT_KEYS = (
    "text",
    "postText",
    "post_text",
    "content",
    "commentary",
    "description",
    "body",
    "caption",
)
LINKEDIN_URL_KEYS = ("url", "postUrl", "post_url", "link", "activityUrl", "linkedinUrl")
LINKEDIN_TITLE_KEYS = ("authorName", "author", "name", "companyName", "company", "title")


@dataclass
class MiningCandidate:
    text: str
    query: str
    source: str
    url: str = ""
    title: str = ""


def parse_generic_icp_file(raw: bytes, filename: str) -> list[dict[str, Any]]:
    name = filename.lower()
    if name.endswith(".csv"):
        text = raw.decode("utf-8", errors="ignore")
        return [dict(row) for row in csv.DictReader(io.StringIO(text))]

    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip().lower() for value in rows[0]]
    parsed: list[dict[str, Any]] = []
    for values in rows[1:]:
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
        if any(value not in (None, "") for value in row.values()):
            parsed.append(row)
    return parsed


def aggregate_mining_context(rows: list[dict[str, Any]]) -> dict[str, Any]:
    def pick(*names: str) -> str:
        for row in rows:
            normalized = {str(key).strip().lower(): value for key, value in row.items()}
            for name in names:
                value = normalized.get(name)
                if value not in (None, ""):
                    return str(value)
        return ""

    def collect(*names: str) -> list[str]:
        values: list[str] = []
        seen: set[str] = set()
        for row in rows:
            normalized = {str(key).strip().lower(): value for key, value in row.items()}
            for name in names:
                raw = normalized.get(name)
                if raw in (None, ""):
                    continue
                for part in re.split(r"[\n;,|]+", str(raw)):
                    item = part.strip()
                    key = item.lower()
                    if item and key not in seen:
                        seen.add(key)
                        values.append(item)
        return values

    return {
        "company_name": pick("company", "company name", "client", "brand") or "Uploaded ICP",
        "website": pick("website", "company website", "url"),
        "services": collect("services", "service", "offer", "service category", "keywords"),
        "geography": pick("geography", "region", "country", "location"),
        "icp": pick("icp", "ideal customer profile", "persona", "target audience"),
        "positioning": pick("positioning", "value proposition", "notes", "gtm positioning"),
        "customer_examples": collect("customer examples", "sample clients", "industries", "industry"),
        "founder_profile": pick("founder", "founder profile"),
        "lead_objective": pick("lead objective", "objective", "goal", "use case", "mining objective", "campaign objective"),
    }


def _clean_html(value: str) -> str:
    value = TAG_RE.sub(" ", value)
    value = html.unescape(value)
    return re.sub(r"\s+", " ", value).strip()


def _open_url(url: str, timeout: int = 8) -> str:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 LeadVaultAI/1.0 public-web-research",
            "Accept": "text/html,application/xhtml+xml",
        },
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        content_type = response.headers.get("content-type", "")
        if "text" not in content_type and "html" not in content_type:
            return ""
        return response.read(750_000).decode("utf-8", errors="ignore")


def _split_setting(value: str) -> list[str]:
    items = []
    seen: set[str] = set()
    for part in re.split(r"[\n,;|]+", value or ""):
        item = part.strip()
        if not item:
            continue
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        items.append(item)
    return items


def _actor_path(actor_id: str) -> str:
    return urllib.parse.quote(actor_id.replace("/", "~"), safe="~")


def _linkedin_search_url(query: str) -> str:
    encoded = urllib.parse.quote_plus(query)
    return f"https://www.linkedin.com/search/results/content/?keywords={encoded}&datePosted=%22past-week%22&sortBy=%22date_posted%22"


def _apify_inputs(query: str, max_posts: int) -> list[dict[str, Any]]:
    url = _linkedin_search_url(query)
    return [
        {"urls": [{"url": url}], "maxItems": max_posts, "maxPosts": max_posts, "proxy": {"useApifyProxy": True}},
        {"startUrls": [{"url": url}], "maxItems": max_posts, "maxPosts": max_posts, "proxy": {"useApifyProxy": True}},
        {"url": url, "query": query, "maxItems": max_posts, "maxPosts": max_posts, "proxy": {"useApifyProxy": True}},
        {"queries": [query], "maxItems": max_posts, "maxPosts": max_posts, "proxy": {"useApifyProxy": True}},
    ]


def _post_json(url: str, payload: dict[str, Any], timeout: int = 90) -> Any:
    request = urllib.request.Request(
        url,
        data=json_dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json", "Accept": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        raw = response.read(2_500_000).decode("utf-8", errors="ignore")
        if not raw:
            return []
        import json

        return json.loads(raw)


def json_dumps(payload: dict[str, Any]) -> str:
    import json

    return json.dumps(payload, ensure_ascii=False)


def _value_from_keys(item: dict[str, Any], keys: tuple[str, ...]) -> str:
    for key in keys:
        value = item.get(key)
        if value not in (None, ""):
            if isinstance(value, dict):
                nested = _value_from_keys(value, keys)
                if nested:
                    return nested
            return str(value).strip()
    for value in item.values():
        if isinstance(value, dict):
            nested = _value_from_keys(value, keys)
            if nested:
                return nested
    return ""


def _apify_item_to_candidate(item: dict[str, Any], query: str, actor_id: str) -> MiningCandidate | None:
    text = _value_from_keys(item, LINKEDIN_TEXT_KEYS)
    url = _value_from_keys(item, LINKEDIN_URL_KEYS)
    title = _value_from_keys(item, LINKEDIN_TITLE_KEYS)
    if not text and title:
        text = title
    text = _clean_html(text)
    if not text or len(text) < 20:
        return None
    return MiningCandidate(
        text=text[:1200],
        query=query,
        source=f"apify:{actor_id}",
        url=url,
        title=title or "LinkedIn post",
    )


def _run_apify_actor(token: str, actor_id: str, query: str, max_posts: int) -> tuple[list[MiningCandidate], str | None]:
    endpoint = f"https://api.apify.com/v2/acts/{_actor_path(actor_id)}/run-sync-get-dataset-items?token={urllib.parse.quote(token)}&timeout=120&memory=1024"
    last_error: str | None = None
    for payload in _apify_inputs(query, max_posts):
        try:
            data = _post_json(endpoint, payload, timeout=150)
            if isinstance(data, dict):
                rows = data.get("items") or data.get("data") or data.get("results") or []
            else:
                rows = data
            if not isinstance(rows, list):
                rows = []
            candidates = [
                candidate
                for item in rows
                if isinstance(item, dict)
                for candidate in [_apify_item_to_candidate(item, query, actor_id)]
                if candidate is not None
            ]
            if candidates:
                return candidates[:max_posts], None
        except Exception as exc:  # noqa: BLE001
            last_error = str(exc)[:240]
    return [], last_error


def _apify_linkedin_candidates(queries: list[str], max_queries: int, max_posts_per_query: int) -> tuple[list[MiningCandidate], list[dict[str, Any]]]:
    tokens = _split_setting(settings.APIFY_API_TOKENS)
    actors = _split_setting(settings.APIFY_LINKEDIN_ACTORS)
    report: list[dict[str, Any]] = []
    candidates: list[MiningCandidate] = []
    if not tokens or not actors:
        return candidates, [{"status": "skipped", "reason": "missing_apify_tokens_or_actors"}]

    token_index = 0
    for query in queries[:max_queries]:
        query_report = {"query": query, "status": "empty", "actor": "", "items": 0, "error": ""}
        for actor_id in actors:
            token = tokens[token_index % len(tokens)]
            token_index += 1
            rows, error = _run_apify_actor(token, actor_id, query, max_posts_per_query)
            query_report = {
                "query": query,
                "status": "ok" if rows else "empty",
                "actor": actor_id,
                "items": len(rows),
                "error": error or "",
            }
            if rows:
                candidates.extend(rows)
                break
        report.append(query_report)
    return candidates, report


def _duckduckgo_candidates(query: str, limit: int) -> list[MiningCandidate]:
    search_url = "https://duckduckgo.com/html/?" + urllib.parse.urlencode({"q": query})
    try:
        page = _open_url(search_url)
    except Exception:
        return []

    candidates: list[MiningCandidate] = []
    for raw_url, raw_title in LINK_RE.findall(page):
        url = urllib.parse.unquote(raw_url)
        title = _clean_html(raw_title)
        if "uddg=" in url:
            parsed = urllib.parse.parse_qs(urllib.parse.urlparse(url).query)
            url = parsed.get("uddg", [url])[0]
        candidates.append(MiningCandidate(text=title, query=query, source="duckduckgo", url=url, title=title))
        if len(candidates) >= limit:
            break
    return candidates


def _page_signal_candidates(seed: MiningCandidate, limit: int = 4) -> list[MiningCandidate]:
    if not seed.url:
        return []
    try:
        body = _clean_html(_open_url(seed.url))
    except Exception:
        return []
    snippets = [match.group(0).strip() for match in BUYER_SENTENCE.finditer(body)]
    return [
        MiningCandidate(text=snippet[:900], query=seed.query, source="public_page", url=seed.url, title=seed.title)
        for snippet in snippets[:limit]
    ]


def _linkedin_search_urls(queries: list[str]) -> list[dict[str, str]]:
    urls = []
    for query in queries[:20]:
        urls.append({
            "query": query,
            "url": _linkedin_search_url(query),
            "source": "linkedin_search_url",
        })
    return urls


def _candidate_to_lead(candidate: MiningCandidate, classification: Any, intent: Any, tenant_id: str, lead_objective: str) -> dict[str, Any]:
    now = datetime.now(timezone.utc).date().isoformat()
    return {
        "Date Added": now,
        "Estimated Deal Value": "TBD",
        "Client Name": candidate.title or "Unknown buyer",
        "Client LinkedIn Profile URL": "",
        "Title": "",
        "Company Name": candidate.title or "Unknown company",
        "Company Website": candidate.url,
        "Industry": "",
        "Region": "",
        "Client Email": "",
        "Client Phone": "",
        "Number of Employees": "",
        "Lead Source": candidate.source,
        "Client Intent Signal": candidate.text,
        "Client Exact Query": candidate.query,
        "Client Query Post URL": candidate.url,
        "Priority": intent.priority,
        "Service Category": intent.signal_type,
        "Outreach Status": "Not Contacted",
        "Ajroni Offer": intent.recommended_play,
        "Notes": intent.next_action,
        "tenant_id": tenant_id,
        "lead_objective": lead_objective,
        "category": classification.category,
        "intent_score": intent.intent_score,
        "buying_stage": intent.buying_stage,
        "explainability_trace": intent.explainability_trace,
    }


def _classify_candidates(candidates: list[MiningCandidate], tenant_id: str, max_results: int, lead_objective: str = "agency_procurement") -> tuple[list[dict[str, Any]], list[dict[str, Any]], int]:
    accepted: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    seen: set[str] = set()
    for candidate in candidates:
        key = f"{candidate.text}|{candidate.url}".lower()
        if key in seen:
            continue
        seen.add(key)
        classification = verify_post(candidate.text, lead_objective=lead_objective)
        intent = analyze_buying_intent(
            candidate.text,
            account={"company": candidate.title, "website": candidate.url, "source": candidate.source},
            verification_category=classification.category,
        )
        if classification.should_export and intent.intent_score >= 60:
            accepted.append(_candidate_to_lead(candidate, classification, intent, tenant_id, lead_objective))
        else:
            rejected.append({
                "tenant_id": tenant_id,
                "lead_objective": lead_objective,
                "text": candidate.text,
                "query": candidate.query,
                "source": candidate.source,
                "url": candidate.url,
                "classification": {
                    "category": classification.category,
                    "confidence": classification.confidence,
                    "rationale": classification.rationale,
                    "should_export": classification.should_export,
                },
                "buying_intent": asdict(intent),
                "rejection_reason": classification.rationale if not classification.should_export else "intent_score_below_threshold",
            })
        if len(accepted) >= max_results:
            break
    return accepted, rejected, len(seen)


def _persist_mining_rows(tenant_id: str, accepted: list[dict[str, Any]], rejected: list[dict[str, Any]], prefix: str) -> None:
    if accepted:
        save_accepted_export({
            "tenant_id": tenant_id,
            "rows": accepted,
            "export_id": f"{prefix}_{tenant_id}_{int(datetime.now(timezone.utc).timestamp())}",
        })
    for row in rejected[:100]:
        save_rejected_audit(row)


def run_leadvault_mining(
    *,
    tenant_id: str,
    company_name: str,
    website: str = "",
    services: list[str] | None = None,
    geography: str = "",
    icp: str = "",
    positioning: str = "",
    customer_examples: list[str] | None = None,
    founder_profile: str = "",
    max_results: int = 25,
    live_web: bool = True,
    mining_mode: str = "hybrid",
    max_apify_queries: int = 10,
    max_posts_per_query: int = 10,
    lead_objective: str = "",
) -> dict[str, Any]:
    spec = build_leadvault_spec(
        tenant_id=tenant_id,
        company_name=company_name,
        website=website,
        services=services or [],
        geography=geography,
        icp=icp,
        positioning=positioning,
        customer_examples=customer_examples or [],
        founder_profile=founder_profile,
        lead_objective=lead_objective,
    )
    queries = list(dict.fromkeys([*spec.google_queries, *spec.buyer_phrases, *spec.linkedin_queries]))
    apify_queries = list(dict.fromkeys(spec.linkedin_queries + spec.buyer_phrases))
    candidates: list[MiningCandidate] = []
    apify_report: list[dict[str, Any]] = []

    mode = (mining_mode or "hybrid").lower()
    if mode in {"apify", "apify_linkedin", "hybrid"}:
        apify_candidates, apify_report = _apify_linkedin_candidates(
            apify_queries,
            max_queries=max(1, min(max_apify_queries, 25)),
            max_posts_per_query=max(1, min(max_posts_per_query, 25)),
        )
        candidates.extend(apify_candidates)

    if live_web and mode in {"public_web", "web", "hybrid"}:
        per_query = max(1, min(4, max_results // max(1, min(len(queries), 8))))
        for query in queries[:8]:
            seeds = _duckduckgo_candidates(query, per_query)
            candidates.extend(seeds)
            for seed in seeds[:2]:
                candidates.extend(_page_signal_candidates(seed, limit=2))
            if len(candidates) >= max_results * 2:
                break

    accepted, rejected, reviewed = _classify_candidates(candidates, tenant_id, max_results, spec.lead_objective)
    _persist_mining_rows(tenant_id, accepted, rejected, "mined")

    return {
        "status": "success",
        "spec": leadvault_spec_as_dict(spec),
        "accepted": accepted,
        "rejected": rejected[:100],
        "apify_report": apify_report,
        "linkedin_search_urls": _linkedin_search_urls(spec.linkedin_queries),
        "summary": {
            "queries_generated": len(queries),
            "apify_queries_attempted": len([item for item in apify_report if item.get("query")]),
            "apify_items": sum(int(item.get("items", 0) or 0) for item in apify_report),
            "candidates_reviewed": reviewed,
            "accepted": len(accepted),
            "rejected": len(rejected),
            "live_web": live_web,
            "mining_mode": mode,
            "max_apify_queries": max_apify_queries,
            "max_posts_per_query": max_posts_per_query,
            "lead_objective": spec.lead_objective,
            "objective_label": spec.objective_label,
        },
    }


def parse_linkedin_capture_rows(rows: list[dict[str, Any]], default_query: str = "linkedin_capture") -> list[MiningCandidate]:
    candidates: list[MiningCandidate] = []
    for row in rows:
        normalized = {str(key).strip().lower(): value for key, value in row.items()}
        text = str(
            normalized.get("post")
            or normalized.get("post text")
            or normalized.get("text")
            or normalized.get("content")
            or normalized.get("client intent signal")
            or ""
        ).strip()
        if not text:
            continue
        candidates.append(
            MiningCandidate(
                text=text,
                query=str(normalized.get("query") or normalized.get("client exact query") or default_query),
                source="linkedin_capture",
                url=str(normalized.get("url") or normalized.get("post url") or normalized.get("client query post url") or ""),
                title=str(normalized.get("author") or normalized.get("company") or normalized.get("company name") or "LinkedIn post"),
            )
        )
    return candidates


def parse_linkedin_capture_text(text: str, default_query: str = "linkedin_capture") -> list[MiningCandidate]:
    blocks = [block.strip() for block in re.split(r"\n\s*---+\s*\n|\n\s*\n\s*\n+", text or "") if block.strip()]
    if not blocks and text.strip():
        blocks = [text.strip()]
    return [
        MiningCandidate(text=block, query=default_query, source="linkedin_paste", title="LinkedIn pasted post")
        for block in blocks
    ]


def run_linkedin_capture_mining(
    *,
    tenant_id: str,
    posts_text: str = "",
    rows: list[dict[str, Any]] | None = None,
    query: str = "linkedin_capture",
    max_results: int = 50,
    lead_objective: str = "agency_procurement",
) -> dict[str, Any]:
    candidates = parse_linkedin_capture_rows(rows or [], default_query=query)
    candidates.extend(parse_linkedin_capture_text(posts_text, default_query=query))
    accepted, rejected, reviewed = _classify_candidates(candidates, tenant_id, max_results, lead_objective)
    _persist_mining_rows(tenant_id, accepted, rejected, "linkedin")
    return {
        "status": "success",
        "accepted": accepted,
        "rejected": rejected[:100],
        "summary": {
            "source": "linkedin_capture",
            "candidates_reviewed": reviewed,
            "accepted": len(accepted),
            "rejected": len(rejected),
            "live_web": False,
            "lead_objective": lead_objective,
        },
    }
